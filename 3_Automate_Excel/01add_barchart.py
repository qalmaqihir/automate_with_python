"""
pip install openpyxl => this lib will help us with excel func in python
"""
from openpyxl import load_workbook

wb = load_workbook('pivot_table.xlsx')
sheet=wb['name_sheet']

# Select the active rows and columns

min_col=wb.active.min_column
max_col=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row


#Printing the rows,cols
print(max_row)
print(min_row)
print(max_col)
print(min_col)

# using the above values lets create a bar chart
from openpyxl.char import BarChart, Reference

barchat=BarChart()

data=Reference(sheet,min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row) # for values the min row, col or max row and col will change, check the excel and change them
category=Reference(sheet,min_col=min_col+1, max_col=min_col, min_row=min_row+1, max_row=max_row) # same here as well

barchat.add_data(data,titles_from_data=True)
barchat.set_categories(category)

sheet.add_chart(barchat,'B12')
barchat.title="title of the bar chart"
barchat.style=5
wb.save("BarChart.xlsx")

























